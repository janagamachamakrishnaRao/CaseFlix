import os
import subprocess
from PIL import Image
from reportlab.pdfgen import canvas
from openpyxl import load_workbook


def convert_to_pdf(file_path):

    ext = os.path.splitext(file_path)[1].lower()
    pdf_path = os.path.splitext(file_path)[0] + ".pdf"

    if ext == ".pdf":
        return file_path

    elif ext == ".docx":
        try:
            result = subprocess.run(
                [
                    "libreoffice",
                    "--headless",
                    "--convert-to", "pdf",
                    "--outdir", os.path.dirname(file_path),
                    file_path
                ],
                timeout=60,
                capture_output=True
            )
            print("STDOUT:", result.stdout)
            print("STDERR:", result.stderr)
            print("PDF exists:", os.path.exists(pdf_path))
            if os.path.exists(pdf_path):
                return pdf_path
            else:
                print("LibreOffice ran but PDF not created")
        except FileNotFoundError:
            print("LibreOffice NOT INSTALLED on this system")
        except Exception as e:
            print(f"LibreOffice failed: {e}")

    # TXT
    elif ext == ".txt":
        c = canvas.Canvas(pdf_path)
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
        y = 800
        for line in lines:
            c.drawString(40, y, line[:100])
            y -= 20
            if y < 40:
                c.showPage()
                y = 800
        c.save()
        return pdf_path

    # XLSX
    elif ext == ".xlsx":
        workbook = load_workbook(file_path)
        sheet = workbook.active
        c = canvas.Canvas(pdf_path)
        y = 800
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join([str(cell) for cell in row if cell])
            c.drawString(40, y, row_text[:120])
            y -= 20
            if y < 40:
                c.showPage()
                y = 800
        c.save()
        return pdf_path

    # IMAGE
    elif ext in [".jpg", ".jpeg", ".png"]:
        image = Image.open(file_path)
        rgb = image.convert("RGB")
        rgb.save(pdf_path)
        return pdf_path

    return file_path